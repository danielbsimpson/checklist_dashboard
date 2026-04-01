"""Quick exploration of the xlsx structure."""
import openpyxl, sys, datetime

out = open("data/_explore_output.txt", "w", encoding="utf-8")

wb = openpyxl.load_workbook("data/Daily Goal tracker.xlsx", data_only=True)
out.write(f"Sheet count: {len(wb.sheetnames)}\n")
for name in wb.sheetnames:
    out.write(f"  {name!r}\n")

# Check the 5-week sheet layout and a recent sheet
for sheet_name in ["Test - 5 weeks", wb.sheetnames[-3]]:
    ws = wb[sheet_name]
    out.write(f"\n\nSheet: {ws.title!r}  rows={ws.max_row}  cols={ws.max_column}\n")
    for r in range(1, min(ws.max_row + 1, 60)):
        vals = []
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(r, c).value
            if isinstance(v, datetime.datetime):
                vals.append(v.strftime("%Y-%m-%d"))
            else:
                vals.append(str(v) if v is not None else "")
        # Skip all-empty rows
        if any(vals):
            out.write(f"  R{r:02d}: | {'|'.join(v[:16].ljust(16) for v in vals)}\n")

out.close()
print("DONE")
