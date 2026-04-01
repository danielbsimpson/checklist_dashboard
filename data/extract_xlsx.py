"""
extract_xlsx.py
---------------
Parse the multi-sheet 'Daily Goal tracker.xlsx' workbook and produce a single
flat CSV (data/historical_goals.csv) with one row per day per goal.

Output columns
--------------
    date        – YYYY-MM-DD
    goal        – the task label as it appears in the spreadsheet
    category    – daily | weekly | monthly
    completed   – 1 or 0
    week_label  – "Week 1", "Week 2", etc.
    sheet       – the original sheet name (month identifier)

The script auto-detects two layout variants:
  • "early" (Nov2022, Dec2022) – data begins in column A (col index 0)
  • "later" (Jan2023 +)        – data is shifted one column right (col B, index 1)

It also handles "Bonus Week" blocks that appear in 5-week months.

Usage
-----
    python data/extract_xlsx.py
    # → writes data/historical_goals.csv
"""

from __future__ import annotations

import datetime as dt
import os
import re
import sys
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = SCRIPT_DIR / "Daily Goal tracker.xlsx"
OUTPUT_PATH = SCRIPT_DIR / "historical_goals.csv"

# Sheets to skip (template / test sheets with no real data)
SKIP_SHEETS = {"Test", "Test - 5 weeks"}

# Rows whose col-A (or col-B) value signals "not a task row"
SKIP_ROW_PREFIXES = {"Total", "Progress bar", "Weekly progress", "Daily Progress", "Weekly Progress", "Day"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_date(val) -> dt.date | None:
    """Try to coerce a cell value to a Python date."""
    if val is None:
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    s = str(val).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _is_bool_true(val) -> bool:
    """Return True if the cell represents a checked / True value."""
    if val is True:
        return True
    if isinstance(val, str) and val.strip().lower() == "true":
        return True
    return False


def _is_bool(val) -> bool:
    """Return True if the cell is a True/False value (not a number, not blank)."""
    if isinstance(val, bool):
        return True
    if isinstance(val, str) and val.strip().lower() in ("true", "false"):
        return True
    return False


def _is_skip_row(label: str) -> bool:
    """Return True if this row label is metadata, not a task."""
    if not label:
        return True
    for prefix in SKIP_ROW_PREFIXES:
        if label.startswith(prefix):
            return True
    return False


# ---------------------------------------------------------------------------
# Core: parse one "band" (= a Week-pair block: left week + right week)
# ---------------------------------------------------------------------------

def _parse_daily_band(ws, header_row: int, col_offset: int) -> list[dict]:
    """
    Parse one pair of daily-goal weeks (left block + right block).

    Parameters
    ----------
    ws          : openpyxl worksheet
    header_row  : the row that contains "Week N" and dates
    col_offset  : 0 for early layout, 1 for later layout

    Returns a list of record dicts.
    """
    records: list[dict] = []

    # --- Left week ---
    left_week_label = ws.cell(header_row, 1 + col_offset).value  # e.g. "Week 1"
    left_dates: list[dt.date | None] = []
    for c in range(2 + col_offset, 9 + col_offset):  # 7 days
        left_dates.append(_to_date(ws.cell(header_row, c).value))

    # --- Right week ---
    right_start_col = 10 + col_offset  # col J (or K in later layout)
    right_week_label = ws.cell(header_row, right_start_col).value
    right_dates: list[dt.date | None] = []
    for c in range(right_start_col + 1, right_start_col + 8):
        right_dates.append(_to_date(ws.cell(header_row, c).value))

    # Task rows start 2 rows below the "Week N" header (skip the "Day" row)
    task_start = header_row + 2

    for r in range(task_start, task_start + 20):  # generous upper bound
        # Left block task name
        left_label = ws.cell(r, 1 + col_offset).value
        if left_label is not None:
            left_label = str(left_label).strip()
        if _is_skip_row(left_label or ""):
            break  # hit "Total" or similar → done with this band

        # Left block values
        if left_label:
            for i, d in enumerate(left_dates):
                val = ws.cell(r, 2 + col_offset + i).value
                if d and _is_bool(val):
                    records.append({
                        "date": d.isoformat(),
                        "goal": left_label,
                        "category": "daily",
                        "completed": 1 if _is_bool_true(val) else 0,
                        "week_label": str(left_week_label or ""),
                    })

        # Right block task name
        right_label = ws.cell(r, right_start_col).value
        if right_label is not None:
            right_label = str(right_label).strip()
        if right_label and not _is_skip_row(right_label):
            for i, d in enumerate(right_dates):
                val = ws.cell(r, right_start_col + 1 + i).value
                if d and _is_bool(val):
                    records.append({
                        "date": d.isoformat(),
                        "goal": right_label,
                        "category": "daily",
                        "completed": 1 if _is_bool_true(val) else 0,
                        "week_label": str(right_week_label or ""),
                    })

    return records


def _parse_weekly_goals(ws, header_row: int, col_offset: int, week_labels: list[str],
                        week_start_dates: list[dt.date | None]) -> list[dict]:
    """
    Parse weekly goal columns sitting to the right of the daily grids.

    Weekly goals in the "early" layout start around col S (index 18) for both
    the upper and lower band.  In the "later" layout they're at col S/T (index 19/20).
    """
    records: list[dict] = []

    # Find the column that says "Weekly Goals" on the header_row + 1 (the "Day" row)
    weekly_col = None
    for c in range(18 + col_offset, 25 + col_offset):
        val = ws.cell(header_row + 1, c).value
        if val and str(val).strip() == "Weekly Goals":
            weekly_col = c
            break
        # Also check the header row itself
        val2 = ws.cell(header_row, c).value
        if val2 and str(val2).strip() == "Weekly Goals":
            weekly_col = c
            break

    if weekly_col is None:
        return records

    # The week value columns are weekly_col+1, weekly_col+2, (maybe weekly_col+3 for Bonus)
    num_week_cols = len(week_labels)

    task_start = header_row + 2
    for r in range(task_start, task_start + 15):
        label = ws.cell(r, weekly_col).value
        if label is not None:
            label = str(label).strip()
        if _is_skip_row(label or ""):
            break

        if label:
            for i in range(num_week_cols):
                val = ws.cell(r, weekly_col + 1 + i).value
                wsd = week_start_dates[i] if i < len(week_start_dates) else None
                if _is_bool(val) and wsd:
                    records.append({
                        "date": wsd.isoformat(),
                        "goal": label,
                        "category": "weekly",
                        "completed": 1 if _is_bool_true(val) else 0,
                        "week_label": week_labels[i] if i < len(week_labels) else "",
                    })

    return records


def _parse_monthly_goals(ws, header_row: int, col_offset: int,
                         month_date: dt.date) -> list[dict]:
    """
    Parse the monthly goal block.  It sits further to the right, usually
    signalled by a cell containing "Monthly Goals".
    """
    records: list[dict] = []

    # Scan a wide range to find "Monthly Goals"
    monthly_col = None
    monthly_row = None
    for r in range(max(1, header_row - 5), header_row + 20):
        for c in range(20 + col_offset, 30 + col_offset):
            val = ws.cell(r, c).value
            if val and str(val).strip() == "Monthly Goals":
                monthly_col = c
                monthly_row = r
                break
        if monthly_col:
            break

    if monthly_col is None:
        return records

    # Tasks start the row after "Monthly Goals"
    for r in range(monthly_row + 1, monthly_row + 10):
        label = ws.cell(r, monthly_col).value
        if label is not None:
            label = str(label).strip()
        if _is_skip_row(label or ""):
            break

        if label:
            val = ws.cell(r, monthly_col + 1).value
            if _is_bool(val):
                records.append({
                    "date": month_date.isoformat(),
                    "goal": label,
                    "category": "monthly",
                    "completed": 1 if _is_bool_true(val) else 0,
                    "week_label": "",
                })

    return records


def _parse_bonus_week(ws, header_row: int, col_offset: int) -> list[dict]:
    """
    Parse the 'Bonus Week' daily block that appears in 5-week months.
    It sits to the right of the weekly goals in the lower band.
    """
    records: list[dict] = []

    # Scan for "Bonus Week" cell
    bonus_col = None
    bonus_row = None
    for r in range(max(1, header_row - 3), header_row + 5):
        for c in range(23 + col_offset, 32 + col_offset):
            val = ws.cell(r, c).value
            if val and str(val).strip() == "Bonus Week":
                bonus_col = c
                bonus_row = r
                break
        if bonus_col:
            break

    if bonus_col is None:
        return records

    # Dates are in the same row as "Bonus Week", columns bonus_col+1 onward
    dates: list[dt.date | None] = []
    for c in range(bonus_col + 1, bonus_col + 8):
        d = _to_date(ws.cell(bonus_row, c).value)
        if d:
            dates.append(d)
        else:
            break  # stop at first empty

    if not dates:
        return records

    # "Day" row is bonus_row + 1, task rows start at bonus_row + 2
    task_start = bonus_row + 2
    for r in range(task_start, task_start + 20):
        label = ws.cell(r, bonus_col).value
        if label is not None:
            label = str(label).strip()
        if _is_skip_row(label or ""):
            break

        if label:
            for i, d in enumerate(dates):
                val = ws.cell(r, bonus_col + 1 + i).value
                if _is_bool(val):
                    records.append({
                        "date": d.isoformat(),
                        "goal": label,
                        "category": "daily",
                        "completed": 1 if _is_bool_true(val) else 0,
                        "week_label": "Bonus",
                    })

    return records


# ---------------------------------------------------------------------------
# Sheet-level parser
# ---------------------------------------------------------------------------

def _detect_layout(ws) -> tuple[int, int]:
    """
    Detect layout variant.

    Returns (col_offset, first_band_row).
      early (Nov2022):  col_offset=0, first_band_row=3
      later (May2023+): col_offset=1, first_band_row=5
    """
    # In the "later" layout, the "Week 1" label is in col B (index 2).
    # In the "early" layout it's in col A (index 1).
    for r in range(1, 10):
        for c in (1, 2):
            val = ws.cell(r, c).value
            if val and str(val).strip().startswith("Week 1"):
                offset = c - 1  # 0 or 1
                return offset, r
    return 1, 5  # default to later layout


def _get_month_date(ws) -> dt.date | None:
    """Extract the month/year date from the header area of the sheet."""
    for r in range(1, 6):
        for c in range(1, 4):
            d = _to_date(ws.cell(r, c).value)
            if d:
                return d.replace(day=1)
    return None


def parse_sheet(ws) -> list[dict]:
    """Parse an entire month sheet and return all records."""
    records: list[dict] = []

    col_offset, band1_row = _detect_layout(ws)
    month_date = _get_month_date(ws)

    # --- Band 1: Weeks 1 & 2 ---
    records.extend(_parse_daily_band(ws, band1_row, col_offset))

    # Collect week start dates for weekly goal parsing
    w1_date = _to_date(ws.cell(band1_row, 2 + col_offset).value)
    w2_col = 10 + col_offset
    w2_date = _to_date(ws.cell(band1_row, w2_col + 1).value)

    # Weekly goals for band 1
    records.extend(_parse_weekly_goals(
        ws, band1_row, col_offset,
        week_labels=["Week 1", "Week 2"],
        week_start_dates=[w1_date, w2_date],
    ))

    # --- Band 2: Weeks 3 & 4 ---
    # The second band starts roughly 17 rows after the first band
    band2_row = None
    for r in range(band1_row + 10, band1_row + 25):
        val = ws.cell(r, 1 + col_offset).value
        if val and str(val).strip().startswith("Week 3"):
            band2_row = r
            break

    if band2_row:
        records.extend(_parse_daily_band(ws, band2_row, col_offset))

        w3_date = _to_date(ws.cell(band2_row, 2 + col_offset).value)
        w4_date = _to_date(ws.cell(band2_row, w2_col + 1).value)

        # Check for Bonus week column in weekly goals
        # In 5-week months, weekly goals header shows "Week 3 | Week 4 | Bonus"
        bonus_week_date = None
        for c in range(18 + col_offset, 26 + col_offset):
            val = ws.cell(band2_row + 1, c).value
            if val and str(val).strip() == "Bonus":
                # The bonus week start date: look for it in the bonus week block
                for c2 in range(23 + col_offset, 32 + col_offset):
                    bval = ws.cell(band2_row - 2, c2).value
                    if bval and str(bval).strip() == "Bonus Week":
                        bonus_week_date = _to_date(ws.cell(band2_row - 2, c2 + 1).value)
                        break
                    bval2 = ws.cell(band2_row - 1, c2).value
                    if bval2 and str(bval2).strip() == "Bonus Week":
                        bonus_week_date = _to_date(ws.cell(band2_row - 1, c2 + 1).value)
                        break
                break

        week_labels_b2 = ["Week 3", "Week 4"]
        week_dates_b2 = [w3_date, w4_date]
        if bonus_week_date:
            week_labels_b2.append("Bonus")
            week_dates_b2.append(bonus_week_date)

        records.extend(_parse_weekly_goals(
            ws, band2_row, col_offset,
            week_labels=week_labels_b2,
            week_start_dates=week_dates_b2,
        ))

        # Parse bonus week daily block (5-week months)
        records.extend(_parse_bonus_week(ws, band2_row, col_offset))

    # --- Monthly goals (appear once per sheet) ---
    if month_date:
        records.extend(_parse_monthly_goals(ws, band1_row, col_offset, month_date))

    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found.")
        sys.exit(1)

    print(f"Loading {XLSX_PATH.name} ...")
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)

    all_records: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"  [SKIP] '{sheet_name}'")
            continue

        ws = wb[sheet_name]
        sheet_records = parse_sheet(ws)

        # Tag every record with the source sheet
        for rec in sheet_records:
            rec["sheet"] = sheet_name

        print(f"  [OK] {sheet_name:20s} -> {len(sheet_records):>5} records")
        all_records.extend(sheet_records)

    # Write CSV
    if not all_records:
        print("No records extracted!")
        sys.exit(1)

    cols = ["date", "goal", "category", "completed", "week_label", "sheet"]
    with open(OUTPUT_PATH, "w", encoding="utf-8", newline="") as f:
        f.write(",".join(cols) + "\n")
        for rec in all_records:
            row = []
            for c in cols:
                val = str(rec.get(c, ""))
                # Escape commas / quotes in CSV
                if "," in val or '"' in val:
                    val = '"' + val.replace('"', '""') + '"'
                row.append(val)
            f.write(",".join(row) + "\n")

    print(f"\n[DONE] Wrote {len(all_records)} records to {OUTPUT_PATH.name}")
    print(f"   Date range: {min(r['date'] for r in all_records)} -> {max(r['date'] for r in all_records)}")

    # Quick summary
    categories = {}
    for r in all_records:
        cat = r["category"]
        categories[cat] = categories.get(cat, 0) + 1
    for cat, count in sorted(categories.items()):
        print(f"   {cat:10s}: {count:>6} rows")


if __name__ == "__main__":
    main()
